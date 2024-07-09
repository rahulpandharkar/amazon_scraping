import os
import time
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import messagebox

def scrape_amazon_product(url):
    print(f"Scraping URL: {url}")
    
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(url)
    
    # Allow time for page to load
    time.sleep(5)
    
    # Extracting Product Information
    product_info = {
        'Item Name': None,
        'Features and Description': None,
        'Brand': None,
        'Price': None,
        'Item Photo': None,
        'Item Dimensions': None,
        'Item Weight': None,
        'Bought Past Month': None,  # New field for Bought Past Month
        'Quantity Sold Past month': None,  # Existing field
        'Number of Reviews': None,
        'Imported from': None,
        'Manufacturing Details': None,
        'URL': url  # Field for URL
    }

    try:
        product_info['Item Name'] = driver.find_element(By.ID, 'productTitle').text
    except Exception as e:
        print(f"Error fetching Item Name: {e}")

    try:
        product_info['Features and Description'] = driver.find_element(By.ID, 'feature-bullets').text
    except Exception as e:
        print(f"Error fetching Features and Description: {e}")

    try:
        product_info['Brand'] = driver.find_element(By.ID, 'bylineInfo').text
    except Exception as e:
        print(f"Error fetching Brand: {e}")

    try:
        product_info['Price'] = driver.find_element(By.ID, 'priceblock_ourprice').text
    except:
        try:
            product_info['Price'] = driver.find_element(By.ID, 'priceblock_dealprice').text
        except:
            try:
                price_whole = driver.find_element(By.XPATH, "//span[@class='a-price-whole']").text
                price_fraction = driver.find_element(By.XPATH, "//span[@class='a-price-fraction']").text
                product_info['Price'] = f"{price_whole}.{price_fraction}"
            except Exception as e:
                print(f"Error fetching Price: {e}")

    try:
        product_info['Item Photo'] = driver.find_element(By.ID, 'landingImage').get_attribute('src')
    except Exception as e:
        print(f"Error fetching Item Photo: {e}")

    try:
        product_info['Item Dimensions'] = driver.find_element(By.XPATH, "//th[contains(text(),'Product Dimensions')]/following-sibling::td").text
    except Exception as e:
        print(f"Error fetching Item Dimensions: {e}")

    try:
        product_info['Item Weight'] = driver.find_element(By.XPATH, "//th[contains(text(),'Item Weight')]/following-sibling::td").text
    except Exception as e:
        print(f"Error fetching Item Weight: {e}")

    try:
        product_info['Bought Past Month'] = driver.find_element(By.XPATH, "//span[contains(text(),'bought in past month')]/../span[1]").text
    except Exception as e:
        print(f"Error fetching Bought Past Month: {e}")

    try:
        product_info['Quantity Sold Past month'] = driver.find_element(By.XPATH, "//th[contains(text(),'Best Sellers Rank')]/following-sibling::td").text
    except Exception as e:
        print(f"Error fetching Quantity Sold Past month: {e}")

    try:
        product_info['Number of Reviews'] = driver.find_element(By.ID, 'acrCustomerReviewText').text
    except Exception as e:
        print(f"Error fetching Number of Reviews: {e}")

    try:
        product_info['Imported from'] = driver.find_element(By.XPATH, "//th[contains(text(),'Country of Origin')]/following-sibling::td").text
    except Exception as e:
        print(f"Error fetching Imported from: {e}")

    try:
        product_info['Manufacturing Details'] = driver.find_element(By.XPATH, "//th[contains(text(),'Manufacturer')]/following-sibling::td").text
    except Exception as e:
        print(f"Error fetching Manufacturing Details: {e}")

    driver.quit()

    return product_info

def save_to_excel(data, filename='product_info.xlsx'):
    print(f"Saving data to {filename}")
    headers = ['Sr. No', 'Item Name', 'Features and Description', 'Brand', 'Price', 'Item Photo', 'Item Dimensions', 'Item Weight', 'Bought Past Month', 'Quantity Sold Past month', 'Number of Reviews', 'Imported from', 'Manufacturing Details', 'URL']  # Updated headers
    
    max_attempts = 3
    attempt = 0
    while attempt < max_attempts:
        try:
            # Load existing workbook and sheet
            workbook = load_workbook(filename)
            sheet = workbook.active
            start_row = sheet.max_row + 1
            print(f"Appending data starting from row {start_row}")
            break  # Break out of the loop if successful
        except FileNotFoundError:
            # Create a new workbook and sheet if the file does not exist
            workbook = Workbook()
            sheet = workbook.active
            start_row = 2
            sheet.append(headers)
            print(f"File not found. Created new workbook with headers.")
            break  # Break out of the loop if successful
        except PermissionError:
            attempt += 1
            if attempt < max_attempts:
                print(f"Permission denied. Retrying in 5 seconds... (Attempt {attempt}/{max_attempts})")
                time.sleep(5)
            else:
                raise  # Raise the exception if max attempts reached
    
    # Create directory for product photos if it doesn't exist
    photos_dir = 'Product Pictures'
    if not os.path.exists(photos_dir):
        os.makedirs(photos_dir)
    
    for i, product_info in enumerate(data, start=start_row - 1):
        row = [i] + [product_info[header] for header in headers[1:]]
        sheet.append(row)
        
        # Save product photo
        try:
            img_url = product_info['Item Photo']
            img_data = requests.get(img_url).content
            img_extension = img_url.split('.')[-1]  # Get extension from URL
            img_filename = f"{i}.{img_extension}"  # Use SR number for image filename
            img_path = os.path.join(photos_dir, img_filename)
            
            with open(img_path, 'wb') as img_file:
                img_file.write(img_data)
            
        except Exception as e:
            print(f"Error saving image for row {i}: {e}")
    
    workbook.save(filename)
    print(f"Data saved to {filename}")

def scrape_and_save():
    url = url_entry.get()
    if not url:
        messagebox.showerror("Error", "Please enter a valid URL.")
        return
    
    try:
        product_info = scrape_amazon_product(url)
        save_to_excel([product_info]) # Saved to Excel Sheet named "product_info"
        messagebox.showinfo("Success", "Data scraped and saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Create GUI window
window = tk.Tk()
window.title("Amazon Scraper")
window.geometry('400x200')

# Create URL input field
url_label = tk.Label(window, text="Enter Amazon Product URL:")
url_label.pack(pady=10)
url_entry = tk.Entry(window, width=50)
url_entry.pack(pady=5)

# Create button to start scraping and saving
scrape_button = tk.Button(window, text="Scrape and Save", command=scrape_and_save)
scrape_button.pack(pady=10)

# Run the main tkinter event loop
window.mainloop()
